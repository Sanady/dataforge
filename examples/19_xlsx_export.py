"""XLSX Export — Generating Excel Spreadsheets.

Real-world scenario: Your QA team needs test data in Excel format for
manual review, or you need to seed a system that imports XLSX files.
DataForge can generate data and export directly to .xlsx files.

This example demonstrates:
- Basic XLSX export with schema.to_excel()
- Convenience method forge.to_excel()
- Custom sheet names
- Large dataset export
- Multiple sheets in sequence

NOTE: This example requires the ``openpyxl`` package.
      Install with: pip install openpyxl
"""

import os
import tempfile

from dataforge import DataForge

forge = DataForge(seed=42)

# Check if openpyxl is available
try:
    import openpyxl  # noqa: F401

    HAS_OPENPYXL = True
except ModuleNotFoundError:
    HAS_OPENPYXL = False
    print("NOTE: openpyxl is not installed. Install with: pip install openpyxl")
    print("This example will show the API but skip actual file creation.\n")

# --- Example 1: Basic XLSX export ----------------------------------------

print("=== Basic XLSX Export ===\n")

schema = forge.schema(
    {
        "first_name": "person.first_name",
        "last_name": "person.last_name",
        "email": "internet.email",
        "city": "address.city",
        "phone": "phone_number",
    }
)

if HAS_OPENPYXL:
    xlsx_path = os.path.join(tempfile.gettempdir(), "dataforge_basic.xlsx")
    rows_written = schema.to_excel(path=xlsx_path, count=20)
    file_size = os.path.getsize(xlsx_path)
    print(f"  Written: {rows_written} rows to {xlsx_path}")
    print(f"  File size: {file_size:,} bytes")

    # Read back and verify
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    print(f"  Sheet name: '{ws.title}'")
    print(f"  Columns: {[cell.value for cell in ws[1]]}")
    print(f"  First data row: {[cell.value for cell in ws[2]]}")
    wb.close()
    os.unlink(xlsx_path)
else:
    print("  schema.to_excel(path='output.xlsx', count=20)")
    print("  → Would write 20 rows to output.xlsx")
print()

# --- Example 2: Convenience method (forge.to_excel) ----------------------

print("=== Convenience Method: forge.to_excel() ===\n")

if HAS_OPENPYXL:
    xlsx_path = os.path.join(tempfile.gettempdir(), "dataforge_convenience.xlsx")
    rows_written = forge.to_excel(
        fields=["full_name", "email", "city", "country"],
        path=xlsx_path,
        count=50,
    )
    file_size = os.path.getsize(xlsx_path)
    print(f"  Written: {rows_written} rows")
    print(f"  File size: {file_size:,} bytes")
    os.unlink(xlsx_path)
else:
    print("  forge.to_excel(")
    print("      fields=['full_name', 'email', 'city', 'country'],")
    print("      path='users.xlsx',")
    print("      count=50,")
    print("  )")
    print("  → Would write 50 rows with full_name, email, city, country columns")
print()

# --- Example 3: Custom sheet names ---------------------------------------

print("=== Custom Sheet Names ===\n")

if HAS_OPENPYXL:
    xlsx_path = os.path.join(tempfile.gettempdir(), "dataforge_custom_sheet.xlsx")
    rows_written = schema.to_excel(
        path=xlsx_path,
        count=10,
        sheet_name="Employee Data",
    )
    wb = openpyxl.load_workbook(xlsx_path)
    print(f"  Sheet names: {wb.sheetnames}")
    wb.close()
    os.unlink(xlsx_path)
else:
    print("  schema.to_excel(path='output.xlsx', count=10, sheet_name='Employee Data')")
    print("  → Creates sheet named 'Employee Data'")
print()

# --- Example 4: Large dataset export -------------------------------------

print("=== Large Dataset Export ===\n")

if HAS_OPENPYXL:
    xlsx_path = os.path.join(tempfile.gettempdir(), "dataforge_large.xlsx")

    large_schema = forge.schema(
        {
            "id": "misc.uuid4",
            "name": "person.full_name",
            "email": "internet.email",
            "company": "company.company_name",
            "city": "address.city",
            "country": "address.country",
            "phone": "phone_number",
            "date": "dt.date",
        }
    )

    import time

    start = time.perf_counter()
    rows_written = large_schema.to_excel(path=xlsx_path, count=5_000)
    elapsed = time.perf_counter() - start
    file_size = os.path.getsize(xlsx_path)

    print(f"  Rows: {rows_written:,}")
    print(f"  Time: {elapsed:.2f}s")
    print(f"  File size: {file_size:,} bytes ({file_size / 1024:.0f} KB)")
    print(f"  Rate: {rows_written / elapsed:,.0f} rows/sec")
    os.unlink(xlsx_path)
else:
    print("  large_schema.to_excel(path='large.xlsx', count=5000)")
    print("  → Would write 5,000 rows with 8 columns")
print()

# --- Example 5: Different data types in XLSX ------------------------------

print("=== Various Data Types ===\n")

mixed_schema = forge.schema(
    {
        "uuid": "misc.uuid4",
        "name": "person.full_name",
        "email": "internet.email",
        "date": "dt.date",
        "price": "finance.price",
        "active": "misc.boolean",
        "ip": "internet.ipv4",
        "color": "color.hex_color",
    }
)

if HAS_OPENPYXL:
    xlsx_path = os.path.join(tempfile.gettempdir(), "dataforge_mixed.xlsx")
    mixed_schema.to_excel(path=xlsx_path, count=5)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    print("  Columns and sample values:")
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    for h, v in zip(headers, values):
        print(f"    {h:12s} = {v}")
    wb.close()
    os.unlink(xlsx_path)
else:
    rows = mixed_schema.generate(count=3)
    print("  Sample data (would be in XLSX):")
    for row in rows:
        for k, v in row.items():
            print(f"    {k:12s} = {v}")
        print()
print()

# --- Example 6: Export workflow summary -----------------------------------

print("=== XLSX Export API Summary ===\n")

print("  # Via Schema:")
print("  schema = forge.schema(['full_name', 'email', 'city'])")
print("  schema.to_excel(path='output.xlsx', count=100, sheet_name='Users')\n")

print("  # Via DataForge convenience method:")
print("  forge.to_excel(")
print("      fields=['full_name', 'email', 'city'],")
print("      path='output.xlsx',")
print("      count=100,")
print("      sheet_name='Users',")
print("  )\n")

print("  # Both methods return the number of rows written.")
print("  # Requires: pip install openpyxl")
